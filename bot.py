import logging
import os
import json
from functools import wraps
from dotenv import load_dotenv
from datetime import datetime
from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import (
    Application,
    CommandHandler,
    CallbackQueryHandler,
    ConversationHandler,
    MessageHandler,
    ContextTypes,
    filters,
)
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font

load_dotenv()
ADMIN_ID = int(os.getenv("ADMIN_ID", "0"))

logging.basicConfig(format="%(asctime)s - %(name)s - %(levelname)s - %(message)s", level=logging.INFO)
logger = logging.getLogger(__name__)

WALLET_FILE = "wallet_address.json"
EXCEL_FILE = "shareithub_data_airdrop.xlsx"

INPUT_WALLET_ADDRESS, CHOOSE_WALLET_TYPE, INPUT_CHAIN = range(1, 4)
INPUT_AIRDROP_LINK, INPUT_AIRDROP_TITLE, CHOOSE_AIRDROP_TYPE, CHOOSE_WALLET = range(10, 14)
CHOOSE_WALLET_DELETE = 20
CHOOSE_AIRDROP_DELETE = 30
REMINDER_SETT_MODE, REMINDER_SETT_DELAY, REMINDER_SETT_CHOOSE = range(50, 53)
STOP_REMINDER_CHOOSE = 60

# Konversi dari pixel ke nilai column width openpyxl
PIXEL_WIDTH = 314
COLUMN_WIDTH = (PIXEL_WIDTH - 5) / 7  # ≈ 44.2

def get_workbook():
    """
    Jika file Excel sudah ada, load file tersebut.
    Jika tidak, buat workbook baru dengan header dan format sesuai.
    """
    if os.path.exists(EXCEL_FILE):
        wb = load_workbook(EXCEL_FILE)
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "AirdropData"
        # Header baru: LINK AIRDROP, AIRDROP NAME, AIRDROP TYPE, WALLET ADDRESS, DATE & TIME
        header = ["LINK AIRDROP", "AIRDROP NAME", "AIRDROP TYPE", "WALLET ADDRESS", "DATE & TIME"]
        ws.append(header)
        # Set format header: bold, center alignment
        for col in ws.iter_cols(min_row=1, max_row=1, min_col=1, max_col=5):
            for cell in col:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center")
        # Set lebar kolom untuk kolom A sampai E
        for col in ["A", "B", "C", "D", "E"]:
            ws.column_dimensions[col].width = COLUMN_WIDTH
        wb.save(EXCEL_FILE)
    return wb

def save_workbook(wb):
    wb.save(EXCEL_FILE)

def restricted(func):
    @wraps(func)
    async def wrapped(update: Update, context: ContextTypes.DEFAULT_TYPE, *args, **kwargs):
        user_id = update.effective_user.id
        if user_id != ADMIN_ID:
            if update.message:
                await update.message.reply_text("Maaf, Anda tidak memiliki izin untuk menggunakan bot ini. Jangan lupa Subscribe Channel Youtube & Telegram : SHARE IT HUB")
            elif update.callback_query:
                await update.callback_query.answer("Maaf, Anda tidak memiliki izin.", show_alert=True)
            return
        return await func(update, context, *args, **kwargs)
    return wrapped

def load_wallets():
    if os.path.exists(WALLET_FILE):
        with open(WALLET_FILE, "r") as f:
            try:
                return json.load(f)
            except json.decoder.JSONDecodeError:
                return {}
    return {}

def save_wallet(user_id, address, chain):
    wallets = load_wallets()
    address = address.strip()
    chain = chain.upper()
    if user_id not in wallets:
        wallets[user_id] = []
    wallets[user_id].append({"address": address, "chain": chain})
    with open(WALLET_FILE, "w") as f:
        json.dump(wallets, f, indent=4)

def delete_wallet_by_index(user_id, index: int):
    wallets = load_wallets()
    user_wallets = wallets.get(user_id, [])
    if 0 <= index < len(user_wallets):
        removed = user_wallets.pop(index)
        wallets[user_id] = user_wallets
        with open(WALLET_FILE, "w") as f:
            json.dump(wallets, f, indent=4)
        return removed
    return None

def get_main_keyboard():
    keyboard = [
        [InlineKeyboardButton("✨ Add Airdrop", callback_data="add_airdrop"),
         InlineKeyboardButton("💳 Add Wallet", callback_data="add_wallet")],
        [InlineKeyboardButton("🗑 Delete Wallet", callback_data="delete_wallet"),
         InlineKeyboardButton("📋 List Wallet Address", callback_data="list_wallet")],
        [InlineKeyboardButton("📊 List Airdrop Saved", callback_data="list_airdrop"),
         InlineKeyboardButton("⏰ Reminder List", callback_data="reminder_lst")],
        [InlineKeyboardButton("⚙️ Reminder Sett", callback_data="reminder_sett"),
         InlineKeyboardButton("⏹ Stop Reminder", callback_data="stop_reminder")],
        [InlineKeyboardButton("📥 Download Data", callback_data="download_data")],
        [InlineKeyboardButton("🗑 Delete Airdrop", callback_data="delete_airdrop")],
    ]
    return InlineKeyboardMarkup(keyboard)

@restricted
async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text("This bot was created by: SHARE IT HUB🚀", reply_markup=get_main_keyboard())

async def add_wallet(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    await query.message.reply_text("💳 Silakan masukkan WALLET ADDRESS Anda:")
    return INPUT_WALLET_ADDRESS

async def receive_wallet(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data["wallet_address"] = update.message.text.strip()
    keyboard = [
        [InlineKeyboardButton("EVM", callback_data="wallet_type_evm")],
        [InlineKeyboardButton("Other", callback_data="wallet_type_other")],
    ]
    await update.message.reply_text("Pilih tipe wallet:", reply_markup=InlineKeyboardMarkup(keyboard))
    return CHOOSE_WALLET_TYPE

async def choose_chain(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    if query.data == "wallet_type_evm":
        chain = "EVM"
        save_wallet(str(query.from_user.id), context.user_data["wallet_address"], chain)
        await query.message.reply_text(f"✅ WALLET {context.user_data['wallet_address']} (EVM) BERHASIL DISIMPAN!", reply_markup=get_main_keyboard())
        return ConversationHandler.END
    await query.message.reply_text("Silakan masukkan nama CHAIN untuk wallet Anda:")
    return INPUT_CHAIN

async def save_other_chain(update: Update, context: ContextTypes.DEFAULT_TYPE):
    chain = update.message.text.strip()
    save_wallet(str(update.message.from_user.id), context.user_data["wallet_address"], chain)
    await update.message.reply_text(f"✅ WALLET {context.user_data['wallet_address'].upper()} ({chain.upper()}) BERHASIL DISIMPAN!", reply_markup=get_main_keyboard())
    return ConversationHandler.END

async def add_airdrop(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    await query.message.reply_text("📎 Silakan kirimkan LINK TELEGRAM untuk AIRDROP ini:")
    return INPUT_AIRDROP_LINK

async def receive_airdrop_link(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = update.message.text.strip()
    context.user_data["airdrop_link"] = text.upper()
    await update.message.reply_text("📝 Silakan masukkan JUDUL AIRDROP:")
    return INPUT_AIRDROP_TITLE

async def receive_airdrop_title(update: Update, context: ContextTypes.DEFAULT_TYPE):
    title = update.message.text.strip()
    context.user_data["airdrop_title"] = title.upper()
    keyboard = [
        [InlineKeyboardButton("TESTNET", callback_data="airdrop_type_testnet")],
        [InlineKeyboardButton("AIRDROP", callback_data="airdrop_type_airdrop")],
        [InlineKeyboardButton("NODE", callback_data="airdrop_type_node")],
        [InlineKeyboardButton("OTHER", callback_data="airdrop_type_other")],
    ]
    await update.message.reply_text("Pilih JENIS AIRDROP:", reply_markup=InlineKeyboardMarkup(keyboard))
    return CHOOSE_AIRDROP_TYPE

async def choose_wallet(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    context.user_data["airdrop_type"] = query.data.replace("airdrop_type_", "").upper()
    user_id = str(query.from_user.id)
    wallets = load_wallets().get(user_id, [])
    if not wallets:
        await query.message.reply_text("⚠️ ANDA BELUM MEMILIKI WALLET. SILAKAN TAMBAHKAN WALLET TERLEBIH DAHULU.", reply_markup=get_main_keyboard())
        return ConversationHandler.END
    keyboard = [
        [InlineKeyboardButton(f"{w['address']} ({w['chain']})", callback_data=f"wallet_{i}")]
        for i, w in enumerate(wallets)
    ]
    await query.message.reply_text("PILIH WALLET ADDRESS UNTUK AIRDROP INI:", reply_markup=InlineKeyboardMarkup(keyboard))
    return CHOOSE_WALLET

async def save_airdrop(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    airdrop_link = context.user_data["airdrop_link"]
    airdrop_title = context.user_data["airdrop_title"]
    airdrop_type = context.user_data["airdrop_type"]
    user_id = str(query.from_user.id)
    wallets = load_wallets().get(user_id, [])
    index = int(query.data.replace("wallet_", ""))
    wallet_address = wallets[index]["address"] if index < len(wallets) else "TIDAK DITEMUKAN"
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    new_row = [airdrop_link, airdrop_title, airdrop_type, wallet_address, timestamp]
    
    # Simpan data ke file Excel dan format baris yang baru ditambahkan
    wb = get_workbook()
    ws = wb.active
    ws.append(new_row)
    last_row = ws.max_row
    for col in range(1, 6):
        ws.cell(row=last_row, column=col).alignment = Alignment(horizontal="center", vertical="center")
    save_workbook(wb)
    
    await query.message.reply_text("✅ AIRDROP BERHASIL DISIMPAN KE FILE EXCEL!", reply_markup=get_main_keyboard())
    return ConversationHandler.END

async def delete_wallet(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    user_id = str(query.from_user.id)
    wallets = load_wallets().get(user_id, [])
    if not wallets:
        await query.message.reply_text("⚠️ TIDAK ADA WALLET YANG DITEMUKAN.", reply_markup=get_main_keyboard())
        return ConversationHandler.END
    keyboard = [
        [InlineKeyboardButton(f"❌ {w['address']} ({w['chain']})", callback_data=f"delwallet_{i}")]
        for i, w in enumerate(wallets)
    ]
    await query.message.reply_text("Pilih wallet yang ingin DIHAPUS:", reply_markup=InlineKeyboardMarkup(keyboard))
    return CHOOSE_WALLET_DELETE

async def process_delete_wallet(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    user_id = str(query.from_user.id)
    index_str = query.data.replace("delwallet_", "")
    try:
        index = int(index_str)
    except ValueError:
        await query.message.reply_text("Data tidak valid.", reply_markup=get_main_keyboard())
        return ConversationHandler.END
    removed = delete_wallet_by_index(user_id, index)
    if removed:
        await query.message.reply_text(f"✅ Wallet {removed['address']} ({removed['chain']}) BERHASIL DIHAPUS!", reply_markup=get_main_keyboard())
    else:
        await query.message.reply_text("⚠️ Gagal menghapus wallet.", reply_markup=get_main_keyboard())
    return ConversationHandler.END

async def list_wallet(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    user_id = str(query.from_user.id)
    wallets = load_wallets().get(user_id, [])
    if not wallets:
        await query.message.reply_text("⚠️ Anda belum menyimpan wallet.", reply_markup=get_main_keyboard())
    else:
        text = "💳 *List Wallet Address Anda:*\n"
        for i, w in enumerate(wallets, start=1):
            text += f"{i}. {w['address']} ({w['chain']})\n"
        text = text.replace("_", "\\_")
        await query.message.reply_text(text, parse_mode="Markdown", reply_markup=get_main_keyboard())
    return ConversationHandler.END

async def list_airdrop(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    try:
        wb = get_workbook()
        ws = wb.active
        data = list(ws.values)
        if len(data) < 2:
            await query.message.reply_text("⚠️ Tidak ada data airdrop yang tersimpan.", reply_markup=get_main_keyboard())
            return ConversationHandler.END
        text = "📊 *List Airdrop Saved:*\n\n"
        for idx, row in enumerate(data[1:], start=2):
            if len(row) >= 5:
                text += (
                    f"🎉 *Airdrop Entry #{idx}*\n"
                    f"🔗 *Link:* `{row[0]}`\n"
                    f"📝 *Judul:* **{row[1]}**\n"
                    f"⚙️ *Jenis:* **{row[2]}**\n"
                    f"💼 *Wallet:* `{row[3]}`\n"
                    f"⏰ *Time:* {row[4]}\n"
                    "----------------------------------\n\n"
                )
        await query.message.reply_text(text, parse_mode="Markdown", reply_markup=get_main_keyboard())
    except Exception as e:
        logger.error(e)
        await query.message.reply_text("⚠️ Terjadi kesalahan saat mengambil data.", reply_markup=get_main_keyboard())
    return ConversationHandler.END

async def delete_airdrop(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    try:
        wb = get_workbook()
        ws = wb.active
        data = list(ws.values)
        if len(data) < 2:
            await query.message.reply_text("⚠️ Tidak ada data airdrop yang tersimpan.", reply_markup=get_main_keyboard())
            return ConversationHandler.END
        keyboard = []
        for row_number, row in enumerate(data[1:], start=2):
            if len(row) >= 5:
                button_text = f"{row_number}. {row[1]} - {str(row[0])[:20]}..."
                keyboard.append([InlineKeyboardButton(button_text, callback_data=f"delairdrop_{row_number}")])
        reply_markup = InlineKeyboardMarkup(keyboard)
        await query.message.reply_text("Pilih airdrop yang ingin DIHAPUS:", reply_markup=reply_markup)
        return CHOOSE_AIRDROP_DELETE
    except Exception as e:
        logger.error(e)
        await query.message.reply_text("⚠️ Terjadi kesalahan saat mengambil data.", reply_markup=get_main_keyboard())
        return ConversationHandler.END

async def process_delete_airdrop(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    try:
        row_number = int(query.data.replace("delairdrop_", ""))
        wb = get_workbook()
        ws = wb.active
        max_row = ws.max_row
        if row_number > 1 and row_number <= max_row:
            ws.delete_rows(row_number, 1)
            save_workbook(wb)
            await query.message.reply_text("✅ Airdrop BERHASIL DIHAPUS!", reply_markup=get_main_keyboard())
        else:
            await query.message.reply_text("⚠️ Baris tidak valid.", reply_markup=get_main_keyboard())
    except Exception as e:
        logger.error(e)
        await query.message.reply_text("⚠️ Gagal menghapus airdrop.", reply_markup=get_main_keyboard())
    return ConversationHandler.END

async def reminder_airdrop_job(context: ContextTypes.DEFAULT_TYPE):
    job_data = context.job.data
    chat_id = job_data["chat_id"]
    row_number = job_data["row_number"]
    try:
        wb = get_workbook()
        ws = wb.active
        max_row = ws.max_row
        if row_number >= 2 and row_number <= max_row:
            row = list(ws.values)[row_number - 1]
            text = (
                f"📢 *Reminder Airdrop*\n\n"
                f"🔗 *Link:* `{row[0]}`\n"
                f"📝 *Judul:* **{row[1]}**\n"
                f"⚙️ *Jenis:* **{row[2]}**\n"
                f"💼 *Wallet:* `{row[3]}`\n"
                f"⏰ *Time:* {row[4]}\n"
            )
        else:
            text = "⚠️ Data airdrop tidak ditemukan."
    except Exception as e:
        logger.error(e)
        text = "⚠️ Terjadi kesalahan saat mengambil data airdrop."
    await context.bot.send_message(chat_id=chat_id, text=text, parse_mode="Markdown")

async def reminder_sett(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    keyboard = [
        [InlineKeyboardButton("Auto (4x sehari)", callback_data="rem_sett_mode_auto")],
        [InlineKeyboardButton("Manual (input delay)", callback_data="rem_sett_mode_manual")],
    ]
    await query.message.reply_text("Pilih mode reminder:", reply_markup=InlineKeyboardMarkup(keyboard))
    return REMINDER_SETT_MODE

async def choose_reminder_mode(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    mode = query.data.replace("rem_sett_mode_", "")
    context.user_data["rem_sett_mode"] = mode
    if mode == "auto":
        await query.message.reply_text("Mode auto terpilih. Reminder akan diatur otomatis setiap hari 4x.", reply_markup=get_main_keyboard())
        return await reminder_sett_choose(update, context)
    else:
        await query.message.reply_text("Mode manual terpilih. Silakan masukkan delay dalam menit:")
        return REMINDER_SETT_DELAY

async def reminder_sett_input_delay(update: Update, context: ContextTypes.DEFAULT_TYPE):
    delay = update.message.text.strip()
    context.user_data["reminder_delay"] = delay
    await update.message.reply_text("Delay telah diterima. Sekarang, pilih data airdrop yang ingin di-reminder:")
    return await reminder_sett_choose(update, context)

async def reminder_sett_choose(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query if update.callback_query else None
    chat_id = query.from_user.id if query else update.message.chat_id
    if context.user_data.get("rem_sett_mode") == "auto":
        interval = 21600  # 21600 detik = 6 jam (4x sehari)
        context.user_data["rem_interval"] = interval
    else:
        try:
            delay_minutes = int(context.user_data.get("reminder_delay"))
            interval = delay_minutes * 60
            context.user_data["rem_interval"] = interval
        except (ValueError, TypeError):
            if query:
                await query.message.reply_text("Delay tidak valid.", reply_markup=get_main_keyboard())
            else:
                await update.message.reply_text("Delay tidak valid.", reply_markup=get_main_keyboard())
            return ConversationHandler.END
    try:
        wb = get_workbook()
        ws = wb.active
        data = list(ws.values)
        if len(data) < 2:
            if query:
                await query.message.reply_text("⚠️ Tidak ada data airdrop yang tersimpan.", reply_markup=get_main_keyboard())
            else:
                await update.message.reply_text("⚠️ Tidak ada data airdrop yang tersimpan.", reply_markup=get_main_keyboard())
            return ConversationHandler.END
        keyboard = []
        for row_number, row in enumerate(data[1:], start=2):
            if len(row) >= 5:
                button_text = f"Baris {row_number}: {row[1]}"
                keyboard.append([InlineKeyboardButton(button_text, callback_data=f"rem_sett_choice_{row_number}")])
        reply_markup = InlineKeyboardMarkup(keyboard)
        if query:
            await query.message.reply_text("Pilih data Airdrop yang ingin di-reminder:", reply_markup=reply_markup)
        else:
            await update.message.reply_text("Pilih data Airdrop yang ingin di-reminder:", reply_markup=reply_markup)
        return REMINDER_SETT_CHOOSE
    except Exception as e:
        logger.error(e)
        if query:
            await query.message.reply_text("⚠️ Terjadi kesalahan saat mengambil data.", reply_markup=get_main_keyboard())
        else:
            await update.message.reply_text("⚠️ Terjadi kesalahan saat mengambil data.", reply_markup=get_main_keyboard())
        return ConversationHandler.END

async def reminder_sett_schedule(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    chat_id = query.from_user.id
    row_number = int(query.data.replace("rem_sett_choice_", ""))
    interval = context.user_data.get("rem_interval")
    job_data = {"chat_id": chat_id, "row_number": row_number, "interval": interval}
    job_name = f"{chat_id}_reminder_{row_number}"
    job_queue = context.job_queue if context.job_queue is not None else context.application.job_queue
    if job_queue is None:
        await query.message.reply_text("Job queue tidak tersedia.", reply_markup=get_main_keyboard())
        return ConversationHandler.END
    job = job_queue.run_repeating(reminder_airdrop_job, interval=interval, first=0, data=job_data, name=job_name)
    context.bot_data[job_name] = job
    await query.message.reply_text(f"✅ Reminder untuk data airdrop baris {row_number} telah dijadwalkan.", reply_markup=get_main_keyboard())
    return ConversationHandler.END

async def reminder_list(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    chat_id = query.from_user.id
    job_names = [name for name in context.bot_data.keys() if name.startswith(f"{chat_id}_reminder_")]
    if not job_names:
        await query.message.reply_text("Tidak ada reminder airdrop yang dijadwalkan.", reply_markup=get_main_keyboard())
        return ConversationHandler.END
    text = "⚙️ *Daftar Reminder Airdrop:*\n\n"
    for name in job_names:
        job = context.bot_data.get(name)
        row_number = job.data["row_number"]
        interval_minutes = int(job.data.get("interval", 0) / 60)
        text += f"• Data Airdrop baris {row_number} | Interval: {interval_minutes} menit | Next run: {job.next_run_time}\n"
    await query.message.reply_text(text, parse_mode="Markdown", reply_markup=get_main_keyboard())
    return ConversationHandler.END

async def stop_reminder_menu(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    chat_id = query.from_user.id
    job_names = [name for name in context.bot_data.keys() if name.startswith(f"{chat_id}_reminder_")]
    if not job_names:
        await query.message.reply_text("Tidak ada reminder yang sedang berjalan.", reply_markup=get_main_keyboard())
        return ConversationHandler.END
    keyboard = []
    for name in job_names:
        row_number = name.split("_")[-1]
        keyboard.append([InlineKeyboardButton(f"Stop Reminder Data Airdrop Baris {row_number}", callback_data=f"stoprem_{row_number}")])
    reply_markup = InlineKeyboardMarkup(keyboard)
    await query.message.reply_text("Pilih reminder yang ingin dihentikan:", reply_markup=reply_markup)
    return STOP_REMINDER_CHOOSE

async def process_stop_reminder(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    chat_id = query.from_user.id
    row_number = query.data.replace("stoprem_", "")
    job_name = f"{chat_id}_reminder_{row_number}"
    job = context.bot_data.get(job_name)
    if job:
        job.schedule_removal()
        context.bot_data.pop(job_name, None)
        await query.message.reply_text(f"✅ Reminder untuk data airdrop baris {row_number} telah dihentikan.", reply_markup=get_main_keyboard())
    else:
        await query.message.reply_text("⚠️ Reminder tidak ditemukan.", reply_markup=get_main_keyboard())
    return ConversationHandler.END

async def download_data(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Kirim file Excel ke pengguna."""
    query = update.callback_query
    await query.answer()
    if os.path.exists(EXCEL_FILE):
        await query.message.reply_document(document=open(EXCEL_FILE, "rb"), filename=EXCEL_FILE)
    else:
        await query.message.reply_text("File data tidak ditemukan.", reply_markup=get_main_keyboard())
    return

def main():
    token = os.getenv("TELEGRAM_BOT_TOKEN")
    if not token:
        logger.error("Token bot tidak ditemukan.")
        return

    # Pastikan file Excel tersedia
    get_workbook()

    app = Application.builder().token(token).build()

    app.add_handler(CommandHandler("start", start))
    app.add_handler(CallbackQueryHandler(download_data, pattern="^download_data$"))

    stop_reminder_conv_handler = ConversationHandler(
        entry_points=[CallbackQueryHandler(stop_reminder_menu, pattern="^stop_reminder$")],
        states={
            STOP_REMINDER_CHOOSE: [CallbackQueryHandler(process_stop_reminder, pattern="^stoprem_\\d+$")],
        },
        fallbacks=[],
    )
    app.add_handler(stop_reminder_conv_handler)

    wallet_conv_handler = ConversationHandler(
        entry_points=[CallbackQueryHandler(add_wallet, pattern="^add_wallet$")],
        states={
            INPUT_WALLET_ADDRESS: [MessageHandler(filters.TEXT & ~filters.COMMAND, receive_wallet)],
            CHOOSE_WALLET_TYPE: [CallbackQueryHandler(choose_chain, pattern="^wallet_type_")],
            INPUT_CHAIN: [MessageHandler(filters.TEXT & ~filters.COMMAND, save_other_chain)],
        },
        fallbacks=[],
    )
    app.add_handler(wallet_conv_handler)

    airdrop_conv_handler = ConversationHandler(
        entry_points=[CallbackQueryHandler(add_airdrop, pattern="^add_airdrop$")],
        states={
            INPUT_AIRDROP_LINK: [MessageHandler(filters.TEXT & ~filters.COMMAND, receive_airdrop_link)],
            INPUT_AIRDROP_TITLE: [MessageHandler(filters.TEXT & ~filters.COMMAND, receive_airdrop_title)],
            CHOOSE_AIRDROP_TYPE: [CallbackQueryHandler(choose_wallet, pattern="^airdrop_type_")],
            CHOOSE_WALLET: [CallbackQueryHandler(save_airdrop, pattern="^wallet_")],
        },
        fallbacks=[],
    )
    app.add_handler(airdrop_conv_handler)

    delete_wallet_conv_handler = ConversationHandler(
        entry_points=[CallbackQueryHandler(delete_wallet, pattern="^delete_wallet$")],
        states={
            CHOOSE_WALLET_DELETE: [CallbackQueryHandler(process_delete_wallet, pattern="^delwallet_")],
        },
        fallbacks=[],
    )
    app.add_handler(delete_wallet_conv_handler)

    delete_airdrop_conv_handler = ConversationHandler(
        entry_points=[CallbackQueryHandler(delete_airdrop, pattern="^delete_airdrop$")],
        states={
            CHOOSE_AIRDROP_DELETE: [CallbackQueryHandler(process_delete_airdrop, pattern="^delairdrop_")],
        },
        fallbacks=[],
    )
    app.add_handler(delete_airdrop_conv_handler)

    app.add_handler(CallbackQueryHandler(list_wallet, pattern="^list_wallet$"))
    app.add_handler(CallbackQueryHandler(list_airdrop, pattern="^list_airdrop$"))
    app.add_handler(CallbackQueryHandler(reminder_list, pattern="^reminder_lst$"))

    reminder_sett_conv_handler = ConversationHandler(
        entry_points=[CallbackQueryHandler(reminder_sett, pattern="^reminder_sett$")],
        states={
            REMINDER_SETT_MODE: [CallbackQueryHandler(choose_reminder_mode, pattern="^rem_sett_mode_(auto|manual)$")],
            REMINDER_SETT_DELAY: [MessageHandler(filters.TEXT & ~filters.COMMAND, reminder_sett_input_delay)],
            REMINDER_SETT_CHOOSE: [CallbackQueryHandler(reminder_sett_schedule, pattern="^rem_sett_choice_\\d+$")],
        },
        fallbacks=[],
    )
    app.add_handler(reminder_sett_conv_handler)

    app.run_polling()

if __name__ == "__main__":
    main()
